from msedge.selenium_tools import Edge
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

file = "C:\\Users\\raine\\Desktop\\testexcel.xlsx"
wb = load_workbook(file)
hoja = wb["testpage"]

driver = Edge(executable_path="C:\\driver_edge\\msedgedriver.exe")
driver.get("https://petstore.octoperf.com/actions/Catalog.action")
time.sleep(5)

largo = hoja.max_column

for i in range(0,largo):
	animal = driver.find_elements(By.CSS_SELECTOR, "div[id='QuickLinks'] a")
	animal[i].click()
	time.sleep(5)
	categoria = driver.find_element(By.TAG_NAME, "h2").text
	animalexcel = hoja.cell(1,i+1).value
	if categoria == animalexcel:

		nombresAnimales = driver.find_elements(By.XPATH, "//tr/td[2]")
		
		for j in range(0,len(nombresAnimales)):

			hoja.cell(j+2,i+1).value = nombresAnimales[j].text

wb.save(file)
wb.close()
driver.quit()